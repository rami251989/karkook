import os
import math
import pandas as pd
import streamlit as st
import psycopg2
from psycopg2 import sql
from openpyxl import load_workbook
from dotenv import load_dotenv
from google.cloud import vision
import re
import base64
import cv2
import numpy as np
from PIL import Image
import io
import tempfile
from concurrent.futures import ThreadPoolExecutor, as_completed

# إضافات لازمة للتبويب الذكي
from rapidfuzz import process, fuzz
import time
import openpyxl

# ---- الإعدادات العامة / البيئة ----
load_dotenv()

USERNAME = "admin"
PASSWORD = "Moraqip@123"

st.set_page_config(page_title="المراقب الذكي", layout="wide")

# ============================ أدوات مساعدة للأداء ============================

# ---- إعداد Google Vision من secrets ----
def setup_google_vision():
    try:
        key_b64 = st.secrets["GOOGLE_VISION_KEY_B64"]
        key_bytes = base64.b64decode(key_b64)
        with open("google_vision.json", "wb") as f:
            f.write(key_bytes)
        os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = "google_vision.json"
        return vision.ImageAnnotatorClient()
    except Exception as e:
        st.error(f"❌ لم يتم تحميل مفتاح Google Vision بشكل صحيح: {e}")
        return None

# ---- اتصال قاعدة البيانات ----
def get_conn():
    return psycopg2.connect(
        dbname=os.environ.get("DB_NAME"),
        user=os.environ.get("DB_USER"),
        password=os.environ.get("DB_PASSWORD"),
        host=os.environ.get("DB_HOST"),
        port=os.environ.get("DB_PORT"),
        sslmode=os.environ.get("DB_SSLMODE", "require")
    )

# ---- استعلام سريع لقائمة أرقام (يختار بين ANY أو جدول مؤقت) ----
def fetch_voters_fast(numbers):
    """
    يأخذ قائمة أرقام كـ strings أو أرقام، ويرجع DataFrame بالحقول المطلوبة.
    - إذا القائمة <= 5000 يستخدم = ANY(%s) (سريع وبسيط).
    - إذا أكبر: ينشئ جدول مؤقت tmp_voters ثم JOIN (أسرع بكثير للقوائم الكبيرة).
    """
    if not numbers:
        return pd.DataFrame(columns=[
            "رقم الناخب","الاسم الثلاثي","الجنس","هاتف","رقم العائلة",
            "اسم مركز الاقتراع","رقم مركز الاقتراع",
            "المدينة","رقم مركز التسجيل","اسم مركز التسجيل","تاريخ الميلاد"
        ])

    numbers = [str(x) for x in numbers]
    # إزالة التكرارات مع الحفاظ على الترتيب
    seen = set()
    uniq = [x for x in numbers if not (x in seen or seen.add(x))]

    conn = get_conn()
    conn.autocommit = True
    rows = []
    try:
        with conn.cursor() as cur:
            if len(uniq) <= 5000:
                cur.execute("""
                    SELECT "رقم الناخب","الاسم الثلاثي","الجنس","هاتف","رقم العائلة",
                           "اسم مركز الاقتراع","رقم مركز الاقتراع",
                           "المدينة","رقم مركز التسجيل","اسم مركز التسجيل","تاريخ الميلاد"
                    FROM "karkook"
                    WHERE "رقم الناخب" = ANY(%s)
                """, (uniq,))
                rows = cur.fetchall()
            else:
                # جدول مؤقت + COPY + JOIN
                cur.execute('CREATE TEMP TABLE tmp_voters (id text) ON COMMIT DROP;')
                buf = io.StringIO("\n".join(uniq))
                cur.copy_from(buf, 'tmp_voters', columns=('id',))
                cur.execute("""
                    SELECT k."رقم الناخب", k."الاسم الثلاثي", k."الجنس", k."هاتف", k."رقم العائلة",
                           k."اسم مركز الاقتراع", k."رقم مركز الاقتراع",
                           k."المدينة", k."رقم مركز التسجيل", k."اسم مركز التسجيل", k."تاريخ الميلاد"
                    FROM "karkook" k
                    JOIN tmp_voters t ON k."رقم الناخب" = t.id
                """)
                rows = cur.fetchall()
    finally:
        conn.close()

    cols = ["رقم الناخب","الاسم الثلاثي","الجنس","هاتف","رقم العائلة",
            "اسم مركز الاقتراع","رقم مركز الاقتراع",
            "المدينة","رقم مركز التسجيل","اسم مركز التسجيل","تاريخ الميلاد"]
    return pd.DataFrame(rows, columns=cols)

# ---- OCR سريع: دفعات + تنفيذ متوازي ----
def ocr_batch(client, uploaded_files_batch):
    """
    يأخذ دفعة من UploadedFile ويرجع list[(filename, full_text)]
    يستخدم DOCUMENT_TEXT_DETECTION عبر batch_annotate_images
    """
    requests = []
    name_map = []
    for f in uploaded_files_batch:
        content = f.read()
        image = vision.Image(content=content)
        features = [vision.Feature(type_=vision.Feature.Type.DOCUMENT_TEXT_DETECTION)]
        req = vision.AnnotateImageRequest(image=image, features=features)
        requests.append(req)
        name_map.append(f.name)

    response = client.batch_annotate_images(requests=requests)
    out = []
    for fname, r in zip(name_map, response.responses):
        full_text = r.full_text_annotation.text if r and r.full_text_annotation else ""
        out.append((fname, full_text))
    return out

def ocr_images_fast(client, uploaded_files, batch_size=16, max_workers=4):
    """
    يقسم الصور إلى دفعات وينفّذها بالتوازي.
    """
    if not uploaded_files:
        return []
    # مهم: UploadedFile.read() يحرك المؤشر؛ لا نحتاج الملف لاحقًا، لذا لا مشكلة.
    chunks = [uploaded_files[i:i+batch_size] for i in range(0, len(uploaded_files), batch_size)]
    results = []
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futures = [ex.submit(ocr_batch, client, chunk) for chunk in chunks]
        for f in as_completed(futures):
            results.extend(f.result())
    return results

# ---- دالة تحويل الجنس ----
def map_gender(x):
    try:
        val = int(float(x))
        return "F" if val == 1 else "M"
    except:
        return "M"

# ---- تسجيل الدخول ----
def login():
    st.markdown(
        """
        <style>
        .login-container {
            display: flex;
            justify-content: center;
            align-items: flex-start;
            height: 100vh;
            padding-top: 10vh;
        }
        .login-box {
            background: #ffffff;
            padding: 1.5rem 2rem;
            border-radius: 12px;
            box-shadow: 0px 2px 12px rgba(0,0,0,0.1);
            text-align: center;
            width: 300px;
        }
        .stTextInput>div>div>input {
            text-align: center;
            font-size: 14px;
            height: 35px;
        }
        .stButton button {
            background: linear-gradient(90deg, #4e73df, #1cc88a);
            color: white;
            border-radius: 6px;
            padding: 0.4rem 0.8rem;
            font-size: 14px;
            font-weight: bold;
            transition: 0.2s;
            width: 100%;
        }
        .stButton button:hover {
            background: linear-gradient(90deg, #1cc88a, #4e73df);
            transform: scale(1.02);
        }
        </style>
        """,
        unsafe_allow_html=True
    )

    st.markdown('<div class="login-container"><div class="login-box">', unsafe_allow_html=True)

    st.markdown("### 🔑 تسجيل الدخول")
    u = st.text_input("👤 اسم المستخدم", key="login_user")
    p = st.text_input("🔒 كلمة المرور", type="password", key="login_pass")

    login_btn = st.button("🚀 دخول", key="login_btn")
    if login_btn:
        if u == USERNAME and p == PASSWORD:
            st.session_state.logged_in = True
            st.rerun()
        else:
            st.error("❌ اسم المستخدم أو كلمة المرور غير صحيحة")

    st.markdown('</div></div>', unsafe_allow_html=True)

# ---- تحقق من حالة الجلسة ----
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

if not st.session_state.logged_in:
    login()
    st.stop()

# ========================== الواجهة بعد تسجيل الدخول ==========================
st.title("📊السيد ناظم - البحث في سجلات الناخبين")
st.markdown("سيتم البحث في قواعد البيانات باستخدام الذكاء الاصطناعي 🤖")

# ====== تبويبات ======
tab_single, tab_file, tab_count = st.tabs(
    [
        "🔍 بحث برقم",
        "📂 رفع ملف Excel",
        "📦 عدّ البطاقات",
    ]
)

# ----------------------------------------------------------------------------- #
# 1) 🔍 البحث برقم واحد (تحويل LIKE إلى = لاستغلال الفهرس)
# ----------------------------------------------------------------------------- #
with tab_single:
    st.subheader("🔍 البحث برقم الناخب")
    voter_input = st.text_input("ادخل رقم الناخب:")
    if st.button("بحث"):
        try:
            conn = get_conn()
            query = """
                SELECT "رقم الناخب","الاسم الثلاثي","الجنس","هاتف","رقم العائلة",
                       "اسم مركز الاقتراع","رقم مركز الاقتراع",
                       "المدينة","رقم مركز التسجيل","اسم مركز التسجيل","تاريخ الميلاد"
                FROM "karkook" WHERE "رقم الناخب" = %s
            """
            df = pd.read_sql_query(query, conn, params=(voter_input.strip(),))
            conn.close()

            if not df.empty:
                df = df.rename(columns={
                    "رقم الناخب": "رقم الناخب",
                    "الاسم الثلاثي": "الاسم",
                    "الجنس": "الجنس",
                    "هاتف": "رقم الهاتف",
                    "رقم العائلة": "رقم العائلة",
                    "اسم مركز الاقتراع": "مركز الاقتراع",
                    "رقم مركز الاقتراع": "رقم مركز الاقتراع",
                    "المدينة": "المدينة",
                    "رقم مركز التسجيل": "رقم مركز التسجيل",
                    "اسم مركز التسجيل": "اسم مركز التسجيل",
                    "تاريخ الميلاد": "تاريخ الميلاد"
                })
                df["الجنس"] = df["الجنس"].apply(map_gender)
                st.dataframe(df, use_container_width=True, height=500)
            else:
                st.warning("⚠️ لم يتم العثور على نتائج")
        except Exception as e:
            st.error(f"❌ خطأ: {e}")

# ----------------------------------------------------------------------------- #
# 2) 📂 رفع ملف Excel (باستخدام fetch_voters_fast لإستعلام سريع)
# ----------------------------------------------------------------------------- #
with tab_file:
    st.subheader("📂 البحث باستخدام ملف Excel")
    uploaded_file = st.file_uploader("📤 ارفع ملف (رقم الناخب)", type=["xlsx"])
    if uploaded_file and st.button("🚀 تشغيل البحث"):
        try:
            voters_df = pd.read_excel(uploaded_file, engine="openpyxl")
            voter_col = "رقم الناخب" if "رقم الناخب" in voters_df.columns else "VoterNo"
            voters_list = voters_df[voter_col].astype(str).tolist()

            df = fetch_voters_fast(voters_list)

            if not df.empty:
                df = df.rename(columns={
                    "رقم الناخب": "رقم الناخب",
                    "الاسم الثلاثي": "الاسم",
                    "الجنس": "الجنس",
                    "هاتف": "رقم الهاتف",
                    "رقم العائلة": "رقم العائلة",
                    "اسم مركز الاقتراع": "مركز الاقتراع",
                    "رقم مركز الاقتراع": "رقم مركز الاقتراع",
                    "المدينة": "المدينة",
                    "رقم مركز التسجيل": "رقم مركز التسجيل",
                    "اسم مركز التسجيل": "اسم مركز التسجيل",
                    "تاريخ الميلاد": "تاريخ الميلاد"
                })
                df["الجنس"] = df["الجنس"].apply(map_gender)

                df["رقم المندوب الرئيسي"] = ""
                df["الحالة"] = 0
                df["ملاحظة"] = ""
                df["رقم المحطة"] = 1

                df = df[["رقم الناخب","الاسم","الجنس","رقم الهاتف",
                         "رقم العائلة","مركز الاقتراع","رقم مركز الاقتراع","رقم المحطة",
                         "رقم المندوب الرئيسي","الحالة","ملاحظة"]]

                # ✅ إيجاد الأرقام غير الموجودة
                found_numbers = set(df["رقم الناخب"].astype(str).tolist())
                missing_numbers = [num for num in voters_list if str(num) not in found_numbers]

                # عرض النتائج الموجودة
                st.dataframe(df, use_container_width=True, height=500)

                # ملف النتائج الموجودة
                output_file = "نتائج_البحث.xlsx"
                df.to_excel(output_file, index=False, engine="openpyxl")
                wb = load_workbook(output_file)
                wb.active.sheet_view.rightToLeft = True
                wb.save(output_file)
                with open(output_file, "rb") as f:
                    st.download_button("⬇️ تحميل النتائج", f,
                        file_name="نتائج_البحث.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

                # عرض وتحميل الأرقام غير الموجودة (كـ DataFrame)
                if missing_numbers:
                    st.warning("⚠️ الأرقام التالية لم يتم العثور عليها في قاعدة البيانات:")
                    missing_df = pd.DataFrame(missing_numbers, columns=["الأرقام غير الموجودة"])
                    st.dataframe(missing_df, use_container_width=True, height=300)
                    miss_file = "missing_numbers.xlsx"
                    missing_df.to_excel(miss_file, index=False, engine="openpyxl")
                    with open(miss_file, "rb") as f:
                        st.download_button("⬇️ تحميل الأرقام غير الموجودة", f,
                            file_name="الأرقام_غير_الموجودة.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                else:
                    st.success("✅ لا توجد أرقام مفقودة في الملف.")
            else:
                st.warning("⚠️ لا يوجد نتائج")
        except Exception as e:
            st.error(f"❌ خطأ: {e}")

# ----------------------------------------------------------------------------- #
# 3) 📦 عدّ البطاقات (OCR دفعات) + بحث سريع + قائمة الأرقام غير الموجودة
# ----------------------------------------------------------------------------- #
# -*- coding: utf-8 -*-
# Streamlit page: OCR أرقام 8 خانات + بحث قاعدة البيانات + قائمة غير الموجودة
# المتطلبات: streamlit, google-cloud-vision, sqlalchemy, psycopg2-binary, pandas, openpyxl

import os
import re
import time
import json
import pandas as pd
import streamlit as st
from io import BytesIO

# Google Vision
from google.cloud import vision
from google.oauth2 import service_account

# Excel
from openpyxl import load_workbook

# SQLAlchemy (PostgreSQL via psycopg2)
from sqlalchemy import create_engine, text
from sqlalchemy.exc import OperationalError, SQLAlchemyError

# =========================
# 🔐 الإعدادات (من secrets)
# =========================
# يُفضّل وضع هذه القيم في .streamlit/secrets.toml بالشكل التالي:
# [general]
# DB_HOST = "db-postgresql-fra1-....ondigitalocean.com"
# DB_PORT = 25060
# DB_NAME = "defaultdb"
# DB_USER = "doadmin"
# DB_PASS = "YOUR_DB_PASSWORD"
#
# GOOGLE_APPLICATION_CREDENTIALS_JSON = """
# { ... محتوى JSON لمستخدم الخدمة من Google ... }
# """

DB_HOST = st.secrets.get("DB_HOST", "db-postgresql-fra1-43509-do-user-11749150-0.d.db.ondigitalocean.com")
DB_PORT = int(st.secrets.get("DB_PORT", 25060))
DB_NAME = st.secrets.get("DB_NAME", "defaultdb")
DB_USER = st.secrets.get("DB_USER", "doadmin")
DB_PASS = st.secrets.get("DB_PASS", "YOUR_PASSWORD_HERE")   # غيّرها في secrets

GOOGLE_SA_JSON = st.secrets.get("GOOGLE_APPLICATION_CREDENTIALS_JSON", None)

# ==================================
# ✅ تهيئة Google Vision (مع كاش)
# ==================================
@st.cache_resource(show_spinner=False)
def setup_google_vision():
    """
    يهيّئ عميل Google Vision. يُفضّل تمرير JSON خدمة من st.secrets.
    """
    try:
        if GOOGLE_SA_JSON:
            # من secrets كـ JSON
            if isinstance(GOOGLE_SA_JSON, str):
                info = json.loads(GOOGLE_SA_JSON)
            else:
                info = GOOGLE_SA_JSON
            creds = service_account.Credentials.from_service_account_info(info)
            client = vision.ImageAnnotatorClient(credentials=creds)
        else:
            # الاعتماد على GOOGLE_APPLICATION_CREDENTIALS في البيئة
            client = vision.ImageAnnotatorClient()
        return client
    except Exception as e:
        st.error(f"❌ خطأ في تهيئة Google Vision: {e}")
        return None

# ===========================================
# 🧠 دالة تحويل الجنس (حسب تنسيق قاعدة بياناتك)
# ===========================================
def map_gender(x):
    if x is None:
        return ""
    s = str(x).strip()
    if s in ["M", "m", "ذكر", "Male", "male", "1"]:
        return "ذكر"
    if s in ["F", "f", "أنثى", "Female", "female", "2"]:
        return "أنثى"
    return s  # كما هو إن لم يطابق

# ===============================================
# 🗄️ محرّك قاعدة البيانات مع Pooling و SSL و Retry
# ===============================================
@st.cache_resource(show_spinner=False)
def get_engine():
    """
    PostgreSQL عبر SQLAlchemy + psycopg2
    - SSL إلزامي على DigitalOcean (sslmode=require)
    - connect_timeout صغير لعدم الانتظار الطويل
    - keepalive لمنع موت الاتصال
    - pool_pre_ping لمنع استخدام اتصال ميت
    """
    params = (
        f"sslmode=require&connect_timeout=5"
        f"&keepalives=1&keepalives_idle=30&keepalives_interval=10&keepalives_count=5"
    )
    url = f"postgresql+psycopg2://{DB_USER}:{DB_PASS}@{DB_HOST}:{DB_PORT}/{DB_NAME}?{params}"
    engine = create_engine(
        url,
        pool_size=5,
        max_overflow=5,
        pool_pre_ping=True,
        pool_recycle=1800,
    )
    return engine

def _query_with_retry(query: str, params: dict, tries: int = 3, delay: float = 1.5):
    """
    تنفيذ استعلام مع إعادة المحاولة على أخطاء الشبكة/الاتصال.
    """
    engine = get_engine()
    last_err = None
    for attempt in range(1, tries + 1):
        try:
            with engine.connect() as conn:
                return pd.read_sql_query(text(query), conn, params=params)
        except OperationalError as e:
            last_err = e
            time.sleep(delay * attempt)  # backoff بسيط
        except SQLAlchemyError as e:
            # أخطاء SQL أخرى: أوقف وأظهر الخطأ
            raise e
    raise last_err if last_err else RuntimeError("فشل الاستعلام بعد إعادة المحاولة.")

def fetch_voters_fast(unique_numbers, chunk_size: int = 500):
    """
    يجلب السجلات على دفعات لتجنّب IN الضخم.
    يستخدم CTE VALUES للربط عبر JOIN وهو أسرع/أكثر استقرارًا في كثير من الحالات.
    """
    if not unique_numbers:
        return pd.DataFrame()

    all_chunks = []
    for i in range(0, len(unique_numbers), chunk_size):
        chunk = [str(x) for x in unique_numbers[i:i+chunk_size]]
        # نبني VALUES (:n0),(:n1)...
        values_clause = ",".join([f"(:n{j})" for j in range(len(chunk))])
        bind_params = {f"n{j}": chunk[j] for j in range(len(chunk))}

        query = f"""
        WITH input(n) AS (
            VALUES {values_clause}
        )
        SELECT
            "رقم الناخب","الاسم الثلاثي","الجنس","هاتف","رقم العائلة",
            "اسم مركز الاقتراع","رقم مركز الاقتراع",
            "المدينة","رقم مركز التسجيل","اسم مركز التسجيل","تاريخ الميلاد"
        FROM "Bagdad" t
        INNER JOIN input i ON t."رقم الناخب" = i.n
        """

        df_part = _query_with_retry(query, bind_params)
        all_chunks.append(df_part)

    if not all_chunks:
        return pd.DataFrame()
    return pd.concat(all_chunks, ignore_index=True)

# =====================================
# 🧾 OCR: استخراج أرقام 8 خانات من صور
# =====================================
def extract_numbers_from_images(client, uploaded_files):
    """
    يأخذ قائمة UploadedFile من Streamlit، يعيد:
    - all_numbers: كل الأرقام مع التكرار
    - number_to_files: mapping رقم -> أسماء الملفات التي ظهر فيها
    - details: ملخّص لكل ملف
    """
    all_numbers, number_to_files, details = [], {}, []

    for img in uploaded_files:
        try:
            content = img.read()
            image = vision.Image(content=content)
            response = client.text_detection(image=image)
            texts = response.text_annotations
            full_text = texts[0].description if texts else ""

            # أرقام من 8 خانات فقط
            found_numbers = re.findall(r"\b\d{8}\b", full_text)
            for n in found_numbers:
                n_str = str(n)
                all_numbers.append(n_str)
                number_to_files.setdefault(n_str, set()).add(img.name)

            details.append({
                "اسم الملف": img.name,
                "عدد البطاقات (أرقام 8 خانات)": len(found_numbers),
                "الأرقام المكتشفة (أرقام 8 خانات فقط)": ", ".join(found_numbers) if found_numbers else "لا يوجد"
            })

        except Exception as e:
            st.warning(f"⚠️ خطأ أثناء معالجة صورة {img.name}: {e}")

    return all_numbers, number_to_files, details

# ==================
# 🖥️ واجهة Streamlit
# ==================
st.set_page_config(page_title="عدّ البطاقات + مطابقة القاعدة", layout="wide")

st.title("📦 عدّ البطاقات (OCR) + بحث سريع + قائمة الأرقام غير الموجودة")

# لو عندك تبويبات في تطبيقك الأساسي، استبدل هذا بما يناسبك
tab_count, = st.tabs(["عدّ البطاقات"])
with tab_count:
    st.subheader("📦 عدّ البطاقات (أرقام 8 خانات) — بحث في القاعدة + الأرقام غير الموجودة")

    imgs_count = st.file_uploader(
        "📤 ارفع صور الصفحات (قد تحتوي أكثر من بطاقة)",
        type=["jpg","jpeg","png"],
        accept_multiple_files=True,
        key="ocr_count"
    )

    if imgs_count and st.button("🚀 عدّ البطاقات والبحث"):
        client = setup_google_vision()
        if client is None:
            st.error("❌ خطأ في إعداد Google Vision.")
        else:
            # --------- OCR ----------
            with st.spinner("⏳ جاري استخراج الأرقام من الصور..."):
                all_numbers, number_to_files, details = extract_numbers_from_images(client, imgs_count)

            total_cards = len(all_numbers)
            unique_numbers = sorted(list(set(all_numbers)))

            st.success("✅ تم الاستخراج الأولي للأرقام")

            # ----------------- بحث في قاعدة البيانات عن الأرقام الموجودة -----------------
            found_df = pd.DataFrame()
            missing_list = []
            db_ok = True

            if unique_numbers:
                try:
                    with st.spinner("🔎 جاري مطابقة الأرقام مع قاعدة البيانات..."):
                        # عدّل chunk_size لو عندك آلاف الأرقام
                        found_df = fetch_voters_fast(unique_numbers, chunk_size=500)

                except Exception as e:
                    db_ok = False
                    st.error(
                        "❌ تعذر الاتصال بقاعدة البيانات حاليًا (Timeout/Network). "
                        "سيتم عرض نتائج OCR فقط وقائمة الأرقام غير الموجودة بناءً على الصور."
                    )
                    st.caption(f"تفاصيل فنية: {e}")

                if db_ok and not found_df.empty:
                    found_df = found_df.rename(columns={
                        "رقم الناخب": "رقم الناخب",
                        "الاسم الثلاثي": "الاسم",
                        "الجنس": "الجنس",
                        "هاتف": "رقم الهاتف",
                        "رقم العائلة": "رقم العائلة",
                        "اسم مركز الاقتراع": "مركز الاقتراع",
                        "رقم مركز الاقتراع": "رقم مركز الاقتراع",
                        "المدينة": "المدينة",
                        "رقم مركز التسجيل": "رقم مركز التسجيل",
                        "اسم مركز التسجيل": "اسم مركز التسجيل",
                        "تاريخ الميلاد": "تاريخ الميلاد"
                    })
                    found_df["الجنس"] = found_df["الجنس"].apply(map_gender)

                    # 🧩 أعمدة إضافية للتوافق مع تبويب رفع الملف
                    found_df["رقم المندوب الرئيسي"] = ""
                    found_df["الحالة"] = 0
                    found_df["ملاحظة"] = ""
                    found_df["رقم المحطة"] = 1

                    # ترتيب الأعمدة
                    found_df = found_df[[
                        "رقم الناخب","الاسم","الجنس","رقم الهاتف",
                        "رقم العائلة","مركز الاقتراع","رقم مركز الاقتراع","رقم المحطة",
                        "رقم المندوب الرئيسي","الحالة","ملاحظة"
                    ]]

                # حساب غير الموجودة
                found_numbers_in_db = set(found_df["رقم الناخب"].astype(str).tolist()) if (db_ok and not found_df.empty) else set()
                for n in unique_numbers:
                    if (not db_ok) or (n not in found_numbers_in_db):
                        files = sorted(list(number_to_files.get(n, [])))
                        missing_list.append({"رقم_الناخب": n, "المصدر(الصور)": ", ".join(files)})

            else:
                st.info("ℹ️ لم يتم العثور على أي أرقام مكوّنة من 8 خانات في الصور المرفوعة.")

            # ----------------- عرض النتائج للمستخدم -----------------
            st.markdown("### 📊 ملخص الاستخراج")
            c1, c2, c3 = st.columns(3)
            c1.metric("إجمالي الأرقام (مع التكرار)", total_cards)
            c2.metric("إجمالي الأرقام الفريدة (8 خانات)", len(unique_numbers))
            c3.metric("عدد الصور المرفوعة", len(imgs_count))

            # تفاصيل كل ملف (اختياري للعرض)
            with st.expander("🧾 تفاصيل استخراج كل صورة"):
                if details:
                    details_df = pd.DataFrame(details)
                    st.dataframe(details_df, use_container_width=True)
                else:
                    st.write("لا تفاصيل.")

            st.markdown("### 🔎 بيانات الناخبين (الموجودة في قاعدة البيانات)")
            if not found_df.empty:
                st.dataframe(found_df, use_container_width=True, height=400)
                out_found = "بيانات_الناخبين_الموجودين.xlsx"
                found_df.to_excel(out_found, index=False, engine="openpyxl")

                # RTL
                wb = load_workbook(out_found)
                wb.active.sheet_view.rightToLeft = True
                wb.save(out_found)

                with open(out_found, "rb") as f:
                    st.download_button(
                        "⬇️ تحميل بيانات الناخبين الموجودة",
                        f,
                        file_name="بيانات_الناخبين_الموجودين.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            else:
                st.warning("⚠️ لم يتم العثور على أي مطابقات في القاعدة.")

            st.markdown("### ❌ الأرقام غير الموجودة في القاعدة (مع اسم الصورة)")
            if missing_list:
                missing_df = pd.DataFrame(missing_list)
                st.dataframe(missing_df, use_container_width=True, height=350)

                miss_file = "missing_numbers_with_files.xlsx"
                missing_df.to_excel(miss_file, index=False, engine="openpyxl")
                with open(miss_file, "rb") as f:
                    st.download_button(
                        "⬇️ تحميل الأرقام غير الموجودة مع المصدر",
                        f,
                        file_name="الأرقام_غير_الموجودة_مع_المصدر.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            else:
                if unique_numbers:
                    st.success("✅ لا توجد أرقام مفقودة (كل الأرقام الموجودة تم إيجادها في قاعدة البيانات).")
